import os
from PIL import Image as PILImage
from tkinter import ttk, messagebox
import tkinter as tk
from decimal import Decimal
from tkcalendar import DateEntry
from datetime import datetime, timedelta, date
import calendar
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from util.ayuda import Ayuda

class PrenominaForm(ttk.Frame):
    def __init__(self, parent, db_manager, usuario_actual):
        super().__init__(parent)
        self.db_manager = db_manager
        self.usuario_actual = usuario_actual
        self.sistema_ayuda = Ayuda()

        self.bind_all("<F1>", self.mostrar_ayuda)
        
        self.pack(fill=tk.BOTH, expand=True)
        
        # Verificar permisos
        alcance = self.db_manager.verificar_permiso(
            usuario_actual['id_usuario'], 
            'nomina.prenomina'
        )

        if not alcance or alcance != 'GLOBAL':
            messagebox.showerror("Error", "No tienes los permisos suficientes para ingresar a este módulo.")
            self.destroy()
            return

        # Frame principal
        self.main_frame = ttk.Frame(self, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Titulo
        ttk.Label(self.main_frame, text="Prenómina", font=('Helvetica', 12, 'bold')).pack(anchor='w', pady=(0,5))
        
        # Frame superior para selección de período
        self.periodo_frame = ttk.LabelFrame(self.main_frame, text="Selección de Período", padding="5")
        self.periodo_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Combobox para períodos
        ttk.Label(self.periodo_frame, text="Período:").pack(side=tk.LEFT, padx=5)
        self.periodo_var = tk.StringVar()
        self.periodo_combo = ttk.Combobox(self.periodo_frame, 
                                        textvariable=self.periodo_var,
                                        state="readonly",
                                        width=40)
        self.periodo_combo.pack(side=tk.LEFT, padx=5)
        
        # Botón para cargar período
        ttk.Button(self.periodo_frame, text="Cargar Período",
                  command=self.cargar_prenominas).pack(side=tk.LEFT, padx=5)
        
        # Frame de búsqueda
        search_frame = ttk.Frame(self.main_frame)
        search_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Label(search_frame, text="Buscar: ").pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var, width=30)
        search_entry.pack(side=tk.LEFT)
        
        # Configurar las columnas para el Treeview
        self.columns = [
            ("nombre", "Nombre"),
            ("apellido", "Apellido"),
            ("cedula", "Cédula"),
            ("cargo", "Cargo"),
            ("salario_base", "Salario Base"),
            ("dias_trabajados", "Días Trabajados"),
            ("dias_descanso", "Días Descanso"),
            ("bonificaciones", "Bonificaciones"),
            ("seguro_social", "Seguro Social"),
            ("rpe", "RPE"),
            ("ley_pol_hab", "Ley Pol. Hab."),
            ("inasistencias", "Inasistencias"),
            ("prestamos", "Préstamos"),
            ("total_deducciones", "Total Deducciones"),
            ("total_a_pagar", "Total a Pagar")
        ]
        
        # Crear Treeview
        self.prenominas_tree = ttk.Treeview(self.main_frame, 
                                          columns=[col[0] for col in self.columns],
                                          show='headings')
        
        # Configurar columnas
        for col, text in self.columns:
            self.prenominas_tree.heading(col, text=text)
            self.prenominas_tree.column(col, width=100)
        
        # Scrollbars
        vsb = ttk.Scrollbar(self.main_frame, orient="vertical", 
                           command=self.prenominas_tree.yview)
        hsb = ttk.Scrollbar(self.main_frame, orient="horizontal",
                           command=self.prenominas_tree.xview)
        self.prenominas_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Posicionar Treeview y scrollbars
        self.prenominas_tree.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        vsb.pack(fill=tk.Y, side=tk.RIGHT)
        hsb.pack(fill=tk.X, side=tk.BOTTOM)
        
        # Frame para botones
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(fill=tk.X, padx=5, pady=5)
        
        ttk.Button(button_frame, text="Procesar Nómina",
                  command=self.procesar_nomina).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cerrar",
                  command=self.destroy).pack(side=tk.RIGHT, padx=5)
        
        # Cargar períodos disponibles
        self.cargar_periodos()
        
        # Vincular búsqueda
        self.search_var.trace('w', self.filtrar_prenominas)
        
        # Lista para almacenar las prenóminas calculadas
        self.prenominas = []

    def mostrar_ayuda(self, event=None):
        """Muestra la ayuda contextual del módulo de empleados"""
        self.sistema_ayuda.mostrar_ayuda('prenomina')

    def cargar_periodos(self):
        """Cargar los períodos disponibles en el combobox"""
        periodos = self.db_manager.obtener_periodos()
        periodos_abiertos = [p for p in periodos if p[4] == 'abierto']  # Solo períodos abiertos
        
        if not periodos_abiertos:
            messagebox.showinfo("Información", "No hay períodos abiertos disponibles")
            return
        
        self.periodo_combo['values'] = [
            f"ID: {p[0]} - {p[1]} ({p[2]} - {p[3]})" for p in periodos_abiertos
        ]

    def calcular_dias_periodo(self, fecha_inicio, fecha_fin, tipo_periodo):
        """Calcula los días trabajados y de descanso en el período"""
        dias_totales = (fecha_fin - fecha_inicio).days + 1
        dias_descanso = 0
        
        # Calcular días de descanso (sábados y domingos)
        fecha_actual = fecha_inicio
        while fecha_actual <= fecha_fin:
            if fecha_actual.weekday() >= 5:  # 5=Sábado, 6=Domingo
                dias_descanso += 1
            fecha_actual += timedelta(days=1)
        
        dias_trabajados = dias_totales - dias_descanso
        return dias_trabajados, dias_descanso

    def calcular_salario_base_periodo(self, salario_mensual, tipo_periodo):
        """
        Calcula el salario base según el tipo de período
        """
        salario_mensual = Decimal(str(salario_mensual))
        
        if tipo_periodo == "Quincenal":
            return salario_mensual / Decimal('2')
        elif tipo_periodo == "Mensual":
            return salario_mensual
        elif tipo_periodo == "Semanal":
            return (salario_mensual * Decimal('12')) / Decimal('52')
        else:
            raise ValueError(f"Tipo de período no válido: {tipo_periodo}")

    def calcular_deducciones_periodo(self, salario_mensual, tipo_periodo, empleado_id):
        """Calcula deducciones obteniendo porcentajes de la base de datos"""
        # Obtener deducciones configuradas
        query = """
        SELECT nombre, porcentaje, tipo
        FROM deducciones 
        """
        deducciones = self.db_manager.ejecutar_query(query, dictionary=True)
        
        # Inicializar acumuladores
        seguro_social = Decimal('0')
        rpe = Decimal('0')
        ley_pol_hab = Decimal('0')
        otras_deducciones = Decimal('0')
        
        # Factor de ajuste según período
        if tipo_periodo == "Quincenal":
            factor_ss_rpe = Decimal('2')  
            factor_lph = Decimal('2')     
        elif tipo_periodo == "Mensual":
            factor_ss_rpe = Decimal('4')  
            factor_lph = Decimal('1')     
        elif tipo_periodo == "Semanal":
            factor_ss_rpe = Decimal('1')  
            factor_lph = Decimal('4')     
            
        for deduccion in deducciones:
            porcentaje = Decimal(str(deduccion['porcentaje']))
            nombre = deduccion['nombre'].lower()
            
            if nombre == "seguro social":
                base_semanal = (salario_mensual * Decimal('12') / Decimal('52'))
                seguro_social = base_semanal * porcentaje * factor_ss_rpe
                
            elif nombre == "rpe":
                base_semanal = (salario_mensual * Decimal('12') / Decimal('52'))
                rpe = base_semanal * porcentaje * factor_ss_rpe
                
            elif nombre == "ley de política habitacional":
                if tipo_periodo == "Quincenal":
                    salario_base_quincenal = salario_mensual / Decimal('2')
                    factor1 = (salario_base_quincenal / Decimal('30')) * (Decimal('45') / Decimal('12'))
                    ley_pol_hab = (factor1 + salario_base_quincenal) * porcentaje
                else:
                    base_mensual = salario_mensual
                    if tipo_periodo != "Mensual":
                        base_mensual = base_mensual / factor_lph
                    ley_pol_hab = base_mensual * porcentaje
            else:
                # Manejar otras deducciones configuradas
                if deduccion['tipo'] == 'porcentaje':
                    otras_deducciones += salario_mensual * porcentaje
                else:
                    otras_deducciones += Decimal(str(deduccion['porcentaje']))
                    
        return seguro_social, rpe, ley_pol_hab, otras_deducciones

    def calcular_valor_inasistencias(self, salario_mensual, num_inasistencias, tipo_periodo):
        """
        Calcula el valor de las inasistencias según el tipo de período
        """
        if tipo_periodo == "Mensual":
            valor_dia = salario_mensual / Decimal('30')
        elif tipo_periodo == "Quincenal":
            valor_dia = (salario_mensual / Decimal('30'))
        elif tipo_periodo == "Semanal":
            valor_dia = (salario_mensual / Decimal('30'))
        
        return valor_dia * Decimal(str(num_inasistencias))

    def ajustar_prestamos_periodo(self, monto_prestamo, tipo_periodo):
        """
        Ajusta el monto de los préstamos según el tipo de período
        """
        if tipo_periodo == "Quincenal":
            return monto_prestamo  # El monto base es quincenal
        elif tipo_periodo == "Mensual":
            return monto_prestamo * Decimal('2')  # Dos cuotas quincenales
        elif tipo_periodo == "Semanal":
            return monto_prestamo / Decimal('2')  # Mitad de la cuota quincenal
        
        return monto_prestamo
    

    def calcular_bonificaciones(self, salario_base, num_inasistencias, empleado_id):
        """Calcula bonificaciones según configuración en BD"""
        query = """
        SELECT nombre, porcentaje, tipo, condicion
        FROM bonificaciones
        """
        bonificaciones = self.db_manager.ejecutar_query(query, dictionary=True)
        
        total_bonificaciones = Decimal('0')
        
        for bono in bonificaciones:
            porcentaje = Decimal(str(bono['porcentaje']))
            aplica_bono = True
            
            # Verificar condiciones
            if bono['condicion'] == 'sin_inasistencias' and num_inasistencias > 0:
                aplica_bono = False
                
            if aplica_bono:
                if bono['tipo'] == 'porcentaje':
                    total_bonificaciones += salario_base * porcentaje
                else:
                    total_bonificaciones += Decimal(str(bono['porcentaje']))
                    
        return total_bonificaciones

    def cargar_prenominas(self):
        """Cargar las prenóminas para el período seleccionado"""
        if not self.periodo_var.get():
            messagebox.showwarning("Advertencia", "Por favor seleccione un período")
            return
        
        # Limpiar Treeview
        for item in self.prenominas_tree.get_children():
            self.prenominas_tree.delete(item)
        
        # Obtener ID y tipo del período
        periodo_id = int(self.periodo_var.get().split('-')[0].replace("ID:", "").strip())
        periodo_info = next((p for p in self.db_manager.obtener_periodos() 
                            if p[0] == periodo_id), None)
        
        if not periodo_info:
            messagebox.showerror("Error", "No se pudo obtener la información del período")
            return
        
        tipo_periodo = periodo_info[1]  # Quincenal, Mensual, Semanal
        fecha_inicio = datetime.strptime(periodo_info[2], '%d-%m-%Y').date()
        fecha_fin = datetime.strptime(periodo_info[3], '%d-%m-%Y').date()
        
        # Obtener empleados
        empleados = self.db_manager.ver_empleados()
        self.prenominas = []
        
        if empleados:
            for emp in empleados:
                # Verificar si el empleado está activo
                status = emp[-1] if len(emp) > 7 else 'activo'
                
                # Saltar empleados inactivos
                if status.lower() == 'inactivo':
                    continue
                    
                # Calcular salario base según tipo de período
                salario_mensual = Decimal(str(emp[5]))
                try:
                    salario_base_periodo = self.calcular_salario_base_periodo(
                        salario_mensual, tipo_periodo)
                except ValueError as e:
                    messagebox.showerror("Error", str(e))
                    return
                
                # Calcular días trabajados y de descanso
                dias_trabajados, dias_descanso = self.calcular_dias_periodo(
                    fecha_inicio, fecha_fin, tipo_periodo)
                
                # Obtener inasistencias
                num_inasistencias, _ = self.db_manager.obtener_inasistencias_empleado(
                    emp[0], fecha_inicio, fecha_fin)
                
                # Calcular bonificaciones
                bonificaciones = self.calcular_bonificaciones(salario_base_periodo, num_inasistencias, emp[0])
                
                # Ajustar días trabajados por inasistencias
                dias_trabajados -= num_inasistencias
                
                # Calcular valor de inasistencias según período
                inasistencias_valor = self.calcular_valor_inasistencias(
                    salario_mensual, num_inasistencias, tipo_periodo)
                
                # Obtener y calcular deducciones según período
                deducciones = self.db_manager.obtener_deducciones_empleado(emp[0])
                seguro_social, rpe, ley_pol_hab, otras_deducciones = self.calcular_deducciones_periodo(
                    salario_mensual, tipo_periodo, emp[0])
                
                # Ajustar préstamos al período
                prestamos = self.db_manager.obtener_prestamos_monto_nomina(emp[0])
                prestamos_valor = self.ajustar_prestamos_periodo(prestamos, tipo_periodo)
                
                # Calcular deducciones totales
                total_deducciones = (seguro_social + rpe + ley_pol_hab +
                                   inasistencias_valor + prestamos_valor)
                
                # Calcular total a pagar (incluyendo bonificaciones)
                total_a_pagar = salario_base_periodo + bonificaciones - total_deducciones
                
                prenomina = (
                    emp[1], emp[2], emp[3], emp[4], salario_base_periodo,
                    dias_trabajados, dias_descanso, bonificaciones,
                    seguro_social, rpe, ley_pol_hab, num_inasistencias,
                    prestamos_valor, total_deducciones, total_a_pagar
                )
                
                self.prenominas.append(prenomina)
                
                # Insertar en Treeview con formato
                values = []
                for val in prenomina:
                    if isinstance(val, Decimal):
                        values.append(f"{val:,.2f}")
                    elif isinstance(val, int):
                        values.append(str(val))
                    else:
                        values.append(val)
                
                self.prenominas_tree.insert('', 'end', values=values)

    def filtrar_prenominas(self, *args):
        """Filtrar prenóminas según el texto de búsqueda"""
        search_text = self.search_var.get().lower()
        
        # Limpiar Treeview
        for item in self.prenominas_tree.get_children():
            self.prenominas_tree.delete(item)
        
        # Filtrar y mostrar coincidencias
        for prenomina in self.prenominas:
            if any(str(value).lower().find(search_text) >= 0 for value in prenomina):
                self.prenominas_tree.insert('', 'end', values=[
                    f"{val:,.2f}" if isinstance(val, Decimal) else val for val in prenomina
                ])

    def procesar_nomina(self):
        """Generar PDFs de las nóminas, cerrar período y procesar préstamos"""
        if not self.periodo_var.get():
            messagebox.showwarning("Advertencia", "Por favor seleccione un período")
            return
                
        if not self.prenominas_tree.get_children():
            messagebox.showwarning("Advertencia", "No hay prenóminas para generar PDFs")
            return
                
        # Obtener información del período
        periodo_id = int(self.periodo_var.get().split('-')[0].replace("ID:", "").strip())
        periodo_info = next((p for p in self.db_manager.obtener_periodos() 
                            if p[0] == periodo_id), None)
        
        if not periodo_info:
            messagebox.showerror("Error", "No se pudo obtener la información del período")
            return

        if messagebox.askyesno("Confirmar", 
                            "¿Está seguro de procesar la nómina? Esto cerrará el período actual."):
            try:
                # Crear carpeta para los PDFs
                fecha_inicio = periodo_info[2].strftime("%Y-%m-%d") if isinstance(periodo_info[2], datetime) else periodo_info[2]
                fecha_fin = periodo_info[3].strftime("%Y-%m-%d") if isinstance(periodo_info[3], datetime) else periodo_info[3]
                output_folder = f"nominas_pdf/periodo_{fecha_inicio}-{fecha_fin}"
                if not os.path.exists(output_folder):
                    os.makedirs(output_folder)

                # Obtener items del Treeview
                all_items = self.prenominas_tree.get_children()
                total_empleados = len(all_items)
                empleados_procesados = 0

                # Procesar cada empleado
                for item in all_items:
                    empleado_data = self.prenominas_tree.item(item)['values']
                    
                    try:
                        nombre = str(empleado_data[0])
                        apellido = str(empleado_data[1])
                        cedula = str(empleado_data[2])

                        empleado = self.db_manager.obtener_empleado_por_cedula(cedula)
                        if not empleado:
                            raise ValueError(f"No se encontró empleado con cédula {cedula}")
                        
                        cargo = str(empleado_data[3])
                        sueldo = float(empleado_data[4].replace(',', '')) if isinstance(empleado_data[4], str) else float(empleado_data[4])
                        dias_trabajados = int(empleado_data[5]) if isinstance(empleado_data[5], str) else empleado_data[5]
                        dias_descanso = int(empleado_data[6]) if isinstance(empleado_data[6], str) else empleado_data[6]
                        bonificaciones = float(empleado_data[7].replace(',', '')) if isinstance(empleado_data[7], str) else float(empleado_data[7])
                        seguro_social = float(empleado_data[8].replace(',', '')) if isinstance(empleado_data[8], str) else float(empleado_data[8])
                        rpe = float(empleado_data[9].replace(',', '')) if isinstance(empleado_data[9], str) else float(empleado_data[9])
                        ley_pol_hab = float(empleado_data[10].replace(',', '')) if isinstance(empleado_data[10], str) else float(empleado_data[10])
                        inasistencias = int(empleado_data[11]) if isinstance(empleado_data[11], str) else empleado_data[11]
                        prestamos = float(empleado_data[12].replace(',', '')) if isinstance(empleado_data[12], str) else float(empleado_data[12])
                        
                        # Procesar préstamos del empleado
                        if prestamos > 0:
                            # Obtener el empleado por cédula
                            empleado = self.db_manager.obtener_empleado_por_cedula(cedula)
                            if empleado:
                                # Registrar pago de préstamo
                                self.db_manager.registrar_pago_prestamo({
                                    'id_empleado': empleado[0],
                                    'monto': prestamos,
                                    'fecha': fecha_fin,
                                    'periodo': f"{fecha_inicio} - {fecha_fin}"
                                })

                        # Insertar en tabla nominas
                        datos_nomina = {
                            'id_empleado': empleado[0],
                            'id_periodo': periodo_id,
                            'salario_base': sueldo,
                            'dias_trabajados': dias_trabajados,
                            'dias_descanso': dias_descanso,
                            'bonificaciones': bonificaciones,
                            'seguro_social': seguro_social,
                            'rpe': rpe,
                            'ley_pol_hab': ley_pol_hab,
                            'num_inasistencias': inasistencias,
                            'valor_inasistencias': inasistencias * (sueldo / 15),
                            'prestamos': prestamos,
                            'total_asignaciones': sueldo + bonificaciones,
                            'total_deducciones': seguro_social + rpe + ley_pol_hab + (inasistencias * (sueldo / 15)) + prestamos,
                            'total_pagar': (sueldo + bonificaciones) - (seguro_social + rpe + ley_pol_hab + (inasistencias * (sueldo / 15)) + prestamos),
                            'procesada_por': f"{self.usuario_actual['nombre']} {self.usuario_actual['apellido']}"
                        }

                        self.db_manager.insertar_nomina(datos_nomina)

                        # Generar PDF
                        pdf_filename = os.path.join(output_folder, f"recibo de pago_{cedula}.pdf")
                        self.generar_pdf_nomina(
                            pdf_filename,
                            {
                                'nombre': nombre,
                                'apellido': apellido,
                                'cedula': cedula,
                                'cargo': cargo,
                                'sueldo': sueldo,
                                'seguro_social': seguro_social,
                                'rpe': rpe,
                                'ley_pol_hab': ley_pol_hab,
                                'inasistencias': inasistencias,
                                'prestamos': prestamos,
                                'bonificaciones': bonificaciones,
                                'periodo_info': periodo_info
                            })
                        
                        empleados_procesados += 1

                        # Generar Excel de nómina
                        excel_filename = self.generar_excel_nomina(output_folder, periodo_info, self.prenominas)
                        
                    except (IndexError, ValueError) as e:
                        print(f"Error procesando empleado: {e}")
                        continue

                # Cerrar el período
                self.db_manager.cerrar_periodo(
                    periodo_id,
                    f"{self.usuario_actual['nombre']} {self.usuario_actual['apellido']}",
                    "Cierre regular de período")
                
                # Detalles para la auditoria
                empleados_procesados_str = []
                for item in self.prenominas_tree.get_children():
                    valores = self.prenominas_tree.item(item)['values']
                    empleados_procesados_str.append(
                        f"Empleado: {valores[0]} {valores[1]} - Cédula: {valores[2]} - "
                        f"Total a pagar: {valores[14]}")

                detalle = (
                    f"Procesamiento de nómina:\n"
                    f"Período: {periodo_info[1]} ({periodo_info[2]} - {periodo_info[3]})\n"
                    f"Total de empleados procesados: {len(empleados_procesados_str)}\n\n"
                    f"Detalles:\n" + "\n".join(empleados_procesados_str))
                # Registrar la auditoria
                self.db_manager.registrar_auditoria(
                    usuario=f"{self.usuario_actual['nombre']} {self.usuario_actual['apellido']}",
                    rol=f"{self.usuario_actual['rol']}",
                    accion='Procesó Nómina',
                    tabla='nominas',
                    detalle=detalle)

                messagebox.showinfo(
                    "Proceso Completado",
                    f"Se han generado {empleados_procesados} archivos PDF de nómina\n"
                    f"Se ha generado el archivo Excel consolidado"
                    f"El período ha sido cerrado exitosamente\n"
                    f"Los archivos se encuentran en: {output_folder}")
                
                # Actualizar la interfaz
                self.periodo_var.set('')
                self.cargar_prenominas()
                
            except Exception as e:
                messagebox.showerror(
                    "Error",
                    f"Ocurrió un error durante el proceso: {str(e)}\n"
                    "Por favor contacte al administrador del sistema.")

    def clear_main_frame(self):
        """Limpia el contenido del frame principal"""
        for widget in self.main_frame.winfo_children():
            widget.destroy()

    def generar_pdf_nomina(self, filename, data):
        """Genera un PDF de nómina con cálculos precisos."""
        doc = SimpleDocTemplate(filename, pagesize=letter, 
                                topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []

        styles = getSampleStyleSheet()
        title_style = styles['Title']
        title_style.alignment = 1  # Centrado

        try:
            logo_path = "./imagenes/Logo RHG.jpg"
            img = PILImage.open(logo_path)
            img_width = 1.2*inch
            aspect = img.height / float(img.width)
            logo = Image(logo_path, width=img_width, height=(img_width * aspect))
            
            title_style.fontSize = 14
            header_style = ParagraphStyle(
                'HeaderStyle',
                parent=styles['Normal'],
                fontSize=12,
                alignment=1
            )
            
            header_data = [
                [logo, Paragraph("<b>R.H.G. INVERSIONES, C.A.</b>", header_style)],
                ['', Paragraph("<b>RIF. J-31347671-9</b>", header_style)],
                ['', Paragraph("<b>COMPROBANTE DE PAGO</b>", header_style)]
            ]
            
            header_table = Table(header_data, colWidths=[1.5*inch, 5*inch])
            header_table.setStyle(TableStyle([
                ('ALIGN', (0,0), (0,-1), 'LEFT'),
                ('ALIGN', (1,0), (1,-1), 'CENTER'),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('SPAN', (0,0), (0,2))
            ]))
            
            elements.append(header_table)
            elements.append(Spacer(1, 0.3 * inch))
            
        except Exception as e:
            print(f"Error al cargar el logo: {e}")

        # Calcular días trabajados usando la función existente
        fecha_inicio = datetime.strptime(data['periodo_info'][2], '%d-%m-%Y').date()
        fecha_fin = datetime.strptime(data['periodo_info'][3], '%d-%m-%Y').date()
        tipo_periodo = data['periodo_info'][1]
        dias_laborables, dias_descanso = self.calcular_dias_periodo(fecha_inicio, fecha_fin, tipo_periodo)

        # Bloque de información del período
        periodo_info = [
            [f"Período:", f"{data['periodo_info'][1].upper()}"],
            [f"Del:", f"{data['periodo_info'][2]} al {data['periodo_info'][3]}"]
        ]
        t_periodo = Table(periodo_info, colWidths=[1*inch, 2*inch, 0.7*inch, 2.5*inch])
        t_periodo.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(t_periodo)
        elements.append(Spacer(1, 0.2 * inch))

        # Bloque de datos del trabajador
        sueldo_mensual = data['sueldo'] * 2

        trabajador_info = [
            [f"Apellidos y Nombres:", f"{data['apellido']} {data['nombre']}"],
            [f"Cédula:", f"{data['cedula']}"],
            [f"Salario Mensual:", f"Bs. {sueldo_mensual:,.2f}"],
            [f"Cargo:", f"{data['cargo']}"]
        ]
        t_trabajador = Table(trabajador_info, colWidths=[2.5 * inch, 3.5 * inch])
        t_trabajador.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(t_trabajador)
        elements.append(Spacer(1, 0.2 * inch))

        # Tabla de asignaciones con días trabajados
        valor_dia = data['sueldo'] / 15
        total_dias_trabajados = valor_dia * dias_laborables
        total_dias_descanso = valor_dia * dias_descanso
        total_asignaciones1 = (total_dias_trabajados + total_dias_descanso + data['bonificaciones'])

        asignaciones_data = [
            ["ASIGNACIONES", "Valor", "Días/Porc.", "Monto"],
            ["Días Trabajados", f"Bs. {valor_dia:,.2f}", f"{dias_laborables}", f"Bs. {total_dias_trabajados:,.2f}"],
            ["Días de Descanso", f"Bs. {valor_dia:,.2f}", f"{dias_descanso}", f"Bs. {total_dias_descanso:,.2f}"],
            ["Bono por Asistencia", "-", "5%", f"Bs. {data['bonificaciones']:,.2f}"]
        ]

        # Configuración de la tabla de asignaciones
        t_conceptos = Table(asignaciones_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        t_conceptos.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Alinear nombres a la izquierda
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        # Calcular el valor por día
        valor_dia = data['sueldo'] / 15  # Para quincena dividimos entre 15 días
        monto_inasistencias = valor_dia * data['inasistencias']


        # Tabla de conceptos - Deducciones
        deducciones_data = [
            ["DEDUCCIONES", "Base", "Porc.", "Monto"],
            ["S.S.O.", f"Bs. {data['sueldo']:,.2f}", "4%", f"Bs. {data['seguro_social']:,.2f}"],
            ["R.P.E.", f"Bs. {data['sueldo']:,.2f}", "0.5%", f"Bs. {data['rpe']:,.2f}"],
            ["F.A.O.V.", f"Bs. {data['sueldo']:,.2f}", "1%", f"Bs. {data['ley_pol_hab']:,.2f}"],
        ]

        # Agregar inasistencias y préstamos si existen
        if data['inasistencias'] > 0:
            deducciones_data.append([
                "Inasistencias", 
                f"Bs. {valor_dia:,.2f}",
                f"{data['inasistencias']} día(s)",
                f"Bs. {monto_inasistencias:,.2f}"
            ])
        if data['prestamos'] > 0:
            deducciones_data.append(["Préstamos", "", "", f"Bs. {data['prestamos']:,.2f}"])

        # Calcular totales
        total_asignaciones = data['sueldo'] + data['bonificaciones']
        total_deducciones = (data['seguro_social'] + data['rpe'] + 
                            data['ley_pol_hab'] + monto_inasistencias + 
                            data['prestamos'])
        neto_a_cobrar = total_asignaciones - total_deducciones

        # Totales
        totales_data = [
            ["TOTALES", "", "", ""],
            ["Total Asignaciones", "", f"Bs. {total_asignaciones:,.2f}", ""],
            ["Total Deducciones", "", "", f"Bs. {total_deducciones:,.2f}"],
            ["NETO A COBRAR", "", f"Bs. {neto_a_cobrar:,.2f}", ""]
        ]

        # Combinar todas las tablas
        all_data = []
        all_data.extend(asignaciones_data)
        all_data.append(["", "", "", ""])  # Espacio entre secciones
        all_data.extend(deducciones_data)
        all_data.append(["", "", "", ""])  # Espacio entre secciones
        all_data.extend(totales_data)

        t_conceptos = Table(all_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        t_conceptos.setStyle(TableStyle([
            # Estilo para encabezados de sección
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#B8CCE4')),
            ('BACKGROUND', (0, len(asignaciones_data) + 1), (-1, len(asignaciones_data) + 1), colors.HexColor('#F2DCDB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('TEXTCOLOR', (0, len(asignaciones_data) + 1), (-1, len(asignaciones_data) + 1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, len(asignaciones_data) + 1), (-1, len(asignaciones_data) + 1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            # Estilo para totales
            ('FONTNAME', (0, -4), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E2EFD9')),
        ]))
        elements.append(t_conceptos)
        elements.append(Spacer(1, 0.3 * inch))

        # Pie de firma
        elements.append(Paragraph(
            "Declaro haber recibido conforme el monto que me corresponde para el "
            "período según los conceptos especificados en este recibo de pago.",
            styles['Normal']
        ))
        elements.append(Spacer(1, 0.5 * inch))

        # Tabla de firma
        firma_data = [
            ["_____________________", "_____________________", "_____________________"],
            ["Firma del Empleado", "Cédula", "Fecha"],
        ]
        t_firma = Table(firma_data, colWidths=[2*inch, 2*inch, 2*inch])
        t_firma.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        elements.append(t_firma)

        doc.build(elements)

    def generar_excel_nomina(self, output_folder, periodo_info, prenominas):
        """Genera el archivo Excel con el resumen de la nómina"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Nómina"
        
        # Estilos
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        title_font = Font(name='Arial', size=12, bold=True)
        subtitle_font = Font(name='Arial', size=10, bold=True)

        ws.merge_cells('B1:D1')
        ws.merge_cells('B2:D2')
        ws.merge_cells('B4:E4')
        ws.merge_cells('B5:E5')
        
        # Encabezado
        header_cells = [
            (ws['B1'], "R.H.G. INVERSIONES, C.A.", title_font, 'left'),
            (ws['B2'], "RIF. J-31347671-9", subtitle_font, 'left'),
            (ws['B4'], "NOMINA EMPLEADOS Y DIRECTIVOS", title_font, 'left'),
            (ws['B5'], f"PERIODO: {periodo_info[1].upper()} DEL {periodo_info[2]} AL {periodo_info[3]}", subtitle_font, 'left')
        ]
        
        for cell, value, font, align in header_cells:
            cell.value = value
            cell.font = font
            cell.alignment = Alignment(horizontal=align, vertical='center')
        
        # Agregar línea de separación decorativa
        for col in range(1, 15):  # A hasta N
            cell = ws.cell(row=6, column=col)
            cell.border = Border(bottom=Side(style='medium'))
        
        # Insertar fecha de generación
        ws['A7'] = f"Fecha de Generación: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
        ws['A7'].font = Font(italic=True, size=8)

        # Insertar logo
        try:
            logo = Image("./imagenes/Logo RHG.jpg")
            logo.width = 120
            logo.height = 120
            ws.add_image(logo, "A1")
        except Exception as e:
            print(f"Error al cargar el logo: {e}")
        
        # Encabezados de la tabla
        headers = [
            ("NOMBRE Y APELLIDO", 30),
            ("N° CEDULA", 15),
            ("SUELDO BASE", 15),
            ("DIARIO", 12),
            ("DIAS LAB", 10),
            ("BONIFICACIÓN", 15),
            ("TOTAL ASIG", 15),
            ("IVSS", 12),
            ("RPE (SPF)", 12),
            ("F.A.O.V (LPH)", 12),
            ("INASISTENCIAS", 15),
            ("PRESTAMO", 12),
            ("TOTAL DEDUCCIONES", 18),
            ("TOTAL A PAGAR", 18)
        ]
        
        # Configurar encabezados
        row_num = 9
        for col, (header, width) in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Datos
        total_sueldo_base = 0
        total_bonificaciones = 0
        total_asignaciones = 0
        total_ivss = 0
        total_rpe = 0
        total_faov = 0
        total_inasistencias = 0
        total_prestamos = 0
        total_deducciones = 0
        total_pagar = 0
        
        for prenomina in prenominas:
            row_num += 1
            # Convertir valores a float si son string con formato
            sueldo_base = float(str(prenomina[4]).replace('Bs. ', '').replace(',', ''))
            diario = sueldo_base / 15  # Para quincena
            bonificaciones = float(str(prenomina[7]).replace('Bs. ', '').replace(',', ''))
            total_asig = sueldo_base + bonificaciones
            
            row = [
                f"{prenomina[0]} {prenomina[1]}",  # Nombre y Apellido
                prenomina[2],  # Cédula
                sueldo_base,  # Sueldo Base
                diario,  # Diario
                15,  # Días laborables (quincena)
                bonificaciones,  # Bonificación
                total_asig,  # Total Asignaciones
                float(str(prenomina[8]).replace('Bs. ', '').replace(',', '')),  # IVSS
                float(str(prenomina[9]).replace('Bs. ', '').replace(',', '')),  # RPE
                float(str(prenomina[10]).replace('Bs. ', '').replace(',', '')),  # FAOV
                float(str(prenomina[11])) * diario,  # Inasistencias
                float(str(prenomina[12]).replace('Bs. ', '').replace(',', '')),  # Préstamos
                float(str(prenomina[13]).replace('Bs. ', '').replace(',', '')),  # Total Deducciones
                float(str(prenomina[14]).replace('Bs. ', '').replace(',', ''))   # Total a Pagar
            ]
            
            # Actualizar totales
            total_sueldo_base += row[2]
            total_bonificaciones += row[5]
            total_asignaciones += row[6]
            total_ivss += row[7]
            total_rpe += row[8]
            total_faov += row[9]
            total_inasistencias += row[10]
            total_prestamos += row[11]
            total_deducciones += row[12]
            total_pagar += row[13]
            
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal='right')
                    cell.number_format = '#,##0.00'
                else:
                    cell.alignment = Alignment(horizontal='left')
        
        # Fila de totales
        row_num += 1
        totales = [
            "TOTALES",
            "",
            total_sueldo_base,
            "",
            "",
            total_bonificaciones,
            total_asignaciones,
            total_ivss,
            total_rpe,
            total_faov,
            total_inasistencias,
            total_prestamos,
            total_deducciones,
            total_pagar
        ]
        
        for col, value in enumerate(totales, start=1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.font = Font(bold=True)
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
                cell.number_format = '#,##0.00'
        
        # Total general
        row_num += 3
        ws.cell(row=row_num, column=1, value=f"TOTAL DEL {periodo_info[2]} AL {periodo_info[3]}")
        ws.cell(row=row_num, column=14, value=total_pagar).number_format = '#,##0.00'
        
        # Campos de firma
        row_num += 3
        ws.cell(row=row_num, column=1, value=f"REALIZADO POR: {self.usuario_actual['nombre']} {self.usuario_actual['apellido']}")
        ws.cell(row=row_num, column=8, value="VERIFICADO POR:")
        
        # Guardar archivo
        excel_filename = os.path.join(output_folder, f"nomina_{periodo_info[2]}-{periodo_info[3]}.xlsx")
        wb.save(excel_filename)
        return excel_filename